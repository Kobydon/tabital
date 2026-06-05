from flask_restful import Resource, request
from flask_praetorian import auth_required, current_user
from app.models.user import User
from app.models.transaction import Transaction
from app.models.instalment import InstalmentPlan
from app.models.instalment_payment import InstalmentPayment
from app.extensions import db
from datetime import datetime, timedelta
from sqlalchemy import func, and_, extract
import csv
from io import StringIO, BytesIO
from flask import make_response
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

class AdminReportRevenueResource(Resource):
    @auth_required
    def get(self):
        """Get revenue report"""
        current_admin = current_user()
        
        if current_admin.role != 'admin':
            return {"error": "Unauthorized"}, 403
        
        # Get query parameters
        period = request.args.get('period', 'monthly')  # daily, weekly, monthly, yearly
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        
        if not year:
            year = datetime.now().year
        
        # Calculate revenue by period
        revenue_data = []
        
        if period == 'daily':
            # Get last 30 days
            for i in range(30, 0, -1):
                date = datetime.now() - timedelta(days=i)
                start_date = date.replace(hour=0, minute=0, second=0, microsecond=0)
                end_date = date.replace(hour=23, minute=59, second=59, microsecond=999999)
                
                revenue = db.session.query(func.sum(Transaction.amount * 0.1))\
                    .filter(Transaction.status == 'completed',
                           Transaction.completion_date.between(start_date, end_date)).scalar() or 0
                
                revenue_data.append({
                    "date": date.strftime("%Y-%m-%d"),
                    "label": date.strftime("%d %b"),
                    "revenue": float(revenue),
                    "transactions": Transaction.query.filter(
                        Transaction.status == 'completed',
                        Transaction.completion_date.between(start_date, end_date)
                    ).count()
                })
        
        elif period == 'weekly':
            # Get last 12 weeks
            for i in range(12, 0, -1):
                start_date = datetime.now() - timedelta(weeks=i)
                end_date = start_date + timedelta(days=6)
                
                revenue = db.session.query(func.sum(Transaction.amount * 0.1))\
                    .filter(Transaction.status == 'completed',
                           Transaction.completion_date.between(start_date, end_date)).scalar() or 0
                
                revenue_data.append({
                    "date": start_date.strftime("%Y-%m-%d"),
                    "label": f"Week {i}",
                    "revenue": float(revenue),
                    "transactions": Transaction.query.filter(
                        Transaction.status == 'completed',
                        Transaction.completion_date.between(start_date, end_date)
                    ).count()
                })
        
        elif period == 'monthly':
            # Get all months in year
            for month_num in range(1, 13):
                start_date = datetime(year, month_num, 1)
                if month_num == 12:
                    end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
                else:
                    end_date = datetime(year, month_num + 1, 1) - timedelta(days=1)
                
                revenue = db.session.query(func.sum(Transaction.amount * 0.1))\
                    .filter(Transaction.status == 'completed',
                           Transaction.completion_date.between(start_date, end_date)).scalar() or 0
                
                revenue_data.append({
                    "date": start_date.strftime("%Y-%m"),
                    "label": start_date.strftime("%b"),
                    "revenue": float(revenue),
                    "transactions": Transaction.query.filter(
                        Transaction.status == 'completed',
                        Transaction.completion_date.between(start_date, end_date)
                    ).count()
                })
        
        elif period == 'yearly':
            # Get last 5 years
            current_year = datetime.now().year
            for yr in range(current_year - 4, current_year + 1):
                start_date = datetime(yr, 1, 1)
                end_date = datetime(yr, 12, 31, 23, 59, 59)
                
                revenue = db.session.query(func.sum(Transaction.amount * 0.1))\
                    .filter(Transaction.status == 'completed',
                           Transaction.completion_date.between(start_date, end_date)).scalar() or 0
                
                revenue_data.append({
                    "date": str(yr),
                    "label": str(yr),
                    "revenue": float(revenue),
                    "transactions": Transaction.query.filter(
                        Transaction.status == 'completed',
                        Transaction.completion_date.between(start_date, end_date)
                    ).count()
                })
        
        # Total revenue
        total_revenue = sum(d["revenue"] for d in revenue_data)
        
        # Growth calculation
        current_period_revenue = revenue_data[-1]["revenue"] if revenue_data else 0
        previous_period_revenue = revenue_data[-2]["revenue"] if len(revenue_data) > 1 else 0
        growth = ((current_period_revenue - previous_period_revenue) / previous_period_revenue * 100) if previous_period_revenue > 0 else 0
        
        return {
            "revenue_data": revenue_data,
            "total_revenue": float(total_revenue),
            "growth": round(growth, 1),
            "period": period,
            "year": year
        }, 200


class AdminReportTransactionsResource(Resource):
    @auth_required
    def get(self):
        """Get transaction report"""
        current_admin = current_user()
        
        if current_admin.role != 'admin':
            return {"error": "Unauthorized"}, 403
        
        period = request.args.get('period', 'monthly')
        year = request.args.get('year', type=int) or datetime.now().year
        
        transaction_data = []
        
        if period == 'monthly':
            for month_num in range(1, 13):
                start_date = datetime(year, month_num, 1)
                if month_num == 12:
                    end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
                else:
                    end_date = datetime(year, month_num + 1, 1) - timedelta(days=1)
                
                total = Transaction.query.filter(
                    Transaction.completion_date.between(start_date, end_date)
                ).count()
                
                completed = Transaction.query.filter(
                    Transaction.status == 'completed',
                    Transaction.completion_date.between(start_date, end_date)
                ).count()
                
                pending = Transaction.query.filter(
                    Transaction.status == 'pending',
                    Transaction.created_at.between(start_date, end_date)
                ).count()
                
                failed = Transaction.query.filter(
                    Transaction.status == 'failed',
                    Transaction.created_at.between(start_date, end_date)
                ).count()
                
                transaction_data.append({
                    "month": start_date.strftime("%b"),
                    "total": total,
                    "completed": completed,
                    "pending": pending,
                    "failed": failed,
                    "success_rate": round((completed / total * 100) if total > 0 else 0, 1)
                })
        
        return {
            "transaction_data": transaction_data,
            "total_transactions": sum(d["total"] for d in transaction_data),
            "avg_success_rate": round(sum(d["success_rate"] for d in transaction_data) / len(transaction_data), 1) if transaction_data else 0
        }, 200


class AdminReportCustomersResource(Resource):
    @auth_required
    def get(self):
        """Get customer report"""
        current_admin = current_user()
        
        if current_admin.role != 'admin':
            return {"error": "Unauthorized"}, 403
        
        period = request.args.get('period', 'monthly')
        year = request.args.get('year', type=int) or datetime.now().year
        
        customer_data = []
        
        if period == 'monthly':
            for month_num in range(1, 13):
                start_date = datetime(year, month_num, 1)
                if month_num == 12:
                    end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
                else:
                    end_date = datetime(year, month_num + 1, 1) - timedelta(days=1)
                
                new_customers = User.query.filter(
                    User.role == 'customer',
                    User.created_at.between(start_date, end_date)
                ).count()
                
                active_customers = db.session.query(func.count(func.distinct(InstalmentPlan.customer_id)))\
                    .filter(InstalmentPlan.created_at.between(start_date, end_date)).scalar() or 0
                
                churned = User.query.filter(
                    User.role == 'customer',
                    User.status == 'suspended',
                    User.updated_at.between(start_date, end_date)
                ).count()
                
                customer_data.append({
                    "month": start_date.strftime("%b"),
                    "new": new_customers,
                    "active": active_customers,
                    "churned": churned,
                    "total": User.query.filter(User.role == 'customer', User.created_at <= end_date).count()
                })
        
        return {
            "customer_data": customer_data,
            "total_customers": User.query.filter(User.role == 'customer').count(),
            "avg_active": round(sum(d["active"] for d in customer_data) / len(customer_data), 1) if customer_data else 0
        }, 200


class AdminReportMerchantsResource(Resource):
    @auth_required
    def get(self):
        """Get merchant report"""
        current_admin = current_user()
        
        if current_admin.role != 'admin':
            return {"error": "Unauthorized"}, 403
        
        period = request.args.get('period', 'monthly')
        year = request.args.get('year', type=int) or datetime.now().year
        
        merchant_data = []
        
        if period == 'monthly':
            for month_num in range(1, 13):
                start_date = datetime(year, month_num, 1)
                if month_num == 12:
                    end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
                else:
                    end_date = datetime(year, month_num + 1, 1) - timedelta(days=1)
                
                new_merchants = User.query.filter(
                    User.role == 'merchant',
                    User.created_at.between(start_date, end_date)
                ).count()
                
                active_merchants = User.query.filter(
                    User.role == 'merchant',
                    User.status.in_(['approved', 'active'])
                ).count()
                
                gmv = db.session.query(func.sum(InstalmentPlan.total_amount))\
                    .filter(InstalmentPlan.created_at.between(start_date, end_date)).scalar() or 0
                
                merchant_data.append({
                    "month": start_date.strftime("%b"),
                    "new": new_merchants,
                    "active": active_merchants,
                    "gmv": float(gmv),
                    "total_gmv": float(db.session.query(func.sum(InstalmentPlan.total_amount)).filter(InstalmentPlan.created_at <= end_date).scalar() or 0)
                })
        
        return {
            "merchant_data": merchant_data,
            "total_merchants": User.query.filter(User.role == 'merchant').count(),
            "total_gmv": float(db.session.query(func.sum(InstalmentPlan.total_amount)).scalar() or 0)
        }, 200


class AdminReportInstalmentsResource(Resource):
    @auth_required
    def get(self):
        """Get instalment report"""
        current_admin = current_user()
        
        if current_admin.role != 'admin':
            return {"error": "Unauthorized"}, 403
        
        period = request.args.get('period', 'monthly')
        year = request.args.get('year', type=int) or datetime.now().year
        
        instalment_data = []
        
        if period == 'monthly':
            for month_num in range(1, 13):
                start_date = datetime(year, month_num, 1)
                if month_num == 12:
                    end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
                else:
                    end_date = datetime(year, month_num + 1, 1) - timedelta(days=1)
                
                new_plans = InstalmentPlan.query.filter(
                    InstalmentPlan.created_at.between(start_date, end_date)
                ).count()
                
                completed = InstalmentPlan.query.filter(
                    InstalmentPlan.status == 'completed',
                    InstalmentPlan.completed_at.between(start_date, end_date)
                ).count()
                
                defaulted = InstalmentPlan.query.filter(
                    InstalmentPlan.status == 'defaulted',
                    InstalmentPlan.created_at.between(start_date, end_date)
                ).count()
                
                collection_rate = round((completed / (completed + defaulted) * 100) if (completed + defaulted) > 0 else 100, 1)
                
                instalment_data.append({
                    "month": start_date.strftime("%b"),
                    "new_plans": new_plans,
                    "completed": completed,
                    "defaulted": defaulted,
                    "collection_rate": collection_rate,
                    "active_plans": InstalmentPlan.query.filter(InstalmentPlan.status == 'active', InstalmentPlan.created_at <= end_date).count()
                })
        
        return {
            "instalment_data": instalment_data,
            "total_active_plans": InstalmentPlan.query.filter_by(status='active').count(),
            "avg_collection_rate": round(sum(d["collection_rate"] for d in instalment_data) / len(instalment_data), 1) if instalment_data else 0
        }, 200


class AdminReportDownloadResource(Resource):
    @auth_required
    def get(self):
        """Download report as Excel/CSV"""
        current_admin = current_user()
        
        if current_admin.role != 'admin':
            return {"error": "Unauthorized"}, 403
        
        report_type = request.args.get('type', 'revenue')
        format_type = request.args.get('format', 'excel')
        period = request.args.get('period', 'monthly')
        year = request.args.get('year', type=int) or datetime.now().year
        
        if format_type == 'csv':
            return self.export_csv(report_type, period, year)
        else:
            return self.export_excel(report_type, period, year)
    
    def export_csv(self, report_type, period, year):
        output = StringIO()
        writer = csv.writer(output)
        
        if report_type == 'revenue':
            writer.writerow(['Period', 'Revenue', 'Transactions', 'Growth (%)'])
            
            start_date = datetime(year, 1, 1)
            for month_num in range(1, 13):
                month_start = datetime(year, month_num, 1)
                if month_num == 12:
                    month_end = datetime(year + 1, 1, 1) - timedelta(days=1)
                else:
                    month_end = datetime(year, month_num + 1, 1) - timedelta(days=1)
                
                revenue = db.session.query(func.sum(Transaction.amount * 0.1))\
                    .filter(Transaction.status == 'completed',
                           Transaction.completion_date.between(month_start, month_end)).scalar() or 0
                
                transactions = Transaction.query.filter(
                    Transaction.status == 'completed',
                    Transaction.completion_date.between(month_start, month_end)
                ).count()
                
                writer.writerow([month_start.strftime("%b %Y"), f"{revenue:.2f}", transactions, ""])
        
        elif report_type == 'transactions':
            writer.writerow(['Month', 'Total Transactions', 'Completed', 'Pending', 'Failed', 'Success Rate (%)'])
            
            for month_num in range(1, 13):
                start_date = datetime(year, month_num, 1)
                if month_num == 12:
                    end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
                else:
                    end_date = datetime(year, month_num + 1, 1) - timedelta(days=1)
                
                total = Transaction.query.filter(
                    Transaction.created_at.between(start_date, end_date)
                ).count()
                
                completed = Transaction.query.filter(
                    Transaction.status == 'completed',
                    Transaction.created_at.between(start_date, end_date)
                ).count()
                
                pending = Transaction.query.filter(
                    Transaction.status == 'pending',
                    Transaction.created_at.between(start_date, end_date)
                ).count()
                
                failed = Transaction.query.filter(
                    Transaction.status == 'failed',
                    Transaction.created_at.between(start_date, end_date)
                ).count()
                
                success_rate = round((completed / total * 100) if total > 0 else 0, 1)
                
                writer.writerow([start_date.strftime("%b %Y"), total, completed, pending, failed, success_rate])
        
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename={report_type}_report_{year}.csv'
        return response
    
    def export_excel(self, report_type, period, year):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{report_type.capitalize()} Report"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        if report_type == 'revenue':
            headers = ['Month', 'Revenue (GHS)', 'Transactions', 'Growth (%)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            row = 2
            for month_num in range(1, 13):
                start_date = datetime(year, month_num, 1)
                if month_num == 12:
                    end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
                else:
                    end_date = datetime(year, month_num + 1, 1) - timedelta(days=1)
                
                revenue = db.session.query(func.sum(Transaction.amount * 0.1))\
                    .filter(Transaction.status == 'completed',
                           Transaction.completion_date.between(start_date, end_date)).scalar() or 0
                
                transactions = Transaction.query.filter(
                    Transaction.status == 'completed',
                    Transaction.completion_date.between(start_date, end_date)
                ).count()
                
                ws.cell(row=row, column=1, value=start_date.strftime("%b %Y"))
                ws.cell(row=row, column=2, value=float(revenue))
                ws.cell(row=row, column=3, value=transactions)
                row += 1
            
            # Add total row
            total_row = row
            ws.cell(row=total_row, column=1, value="TOTAL")
            ws.cell(row=total_row, column=2, value=f"=SUM(B2:B{row-1})")
            ws.cell(row=total_row, column=2).font = Font(bold=True)
        
        elif report_type == 'transactions':
            headers = ['Month', 'Total Transactions', 'Completed', 'Pending', 'Failed', 'Success Rate (%)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            row = 2
            for month_num in range(1, 13):
                start_date = datetime(year, month_num, 1)
                if month_num == 12:
                    end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
                else:
                    end_date = datetime(year, month_num + 1, 1) - timedelta(days=1)
                
                total = Transaction.query.filter(
                    Transaction.created_at.between(start_date, end_date)
                ).count()
                
                completed = Transaction.query.filter(
                    Transaction.status == 'completed',
                    Transaction.created_at.between(start_date, end_date)
                ).count()
                
                pending = Transaction.query.filter(
                    Transaction.status == 'pending',
                    Transaction.created_at.between(start_date, end_date)
                ).count()
                
                failed = Transaction.query.filter(
                    Transaction.status == 'failed',
                    Transaction.created_at.between(start_date, end_date)
                ).count()
                
                success_rate = round((completed / total * 100) if total > 0 else 0, 1)
                
                ws.cell(row=row, column=1, value=start_date.strftime("%b %Y"))
                ws.cell(row=row, column=2, value=total)
                ws.cell(row=row, column=3, value=completed)
                ws.cell(row=row, column=4, value=pending)
                ws.cell(row=row, column=5, value=failed)
                ws.cell(row=row, column=6, value=success_rate)
                row += 1
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={report_type}_report_{year}.xlsx'
        return response


class AdminDashboardKPIResource(Resource):
    @auth_required
    def get(self):
        """Get dashboard KPIs for reports"""
        current_admin = current_user()
        
        if current_admin.role != 'admin':
            return {"error": "Unauthorized"}, 403
        
        # Current year metrics
        current_year = datetime.now().year
        current_month = datetime.now().month
        
        # Year-to-date revenue
        ytd_start = datetime(current_year, 1, 1)
        ytd_revenue = db.session.query(func.sum(Transaction.amount * 0.1))\
            .filter(Transaction.status == 'completed',
                   Transaction.completion_date >= ytd_start).scalar() or 0
        
        # Monthly revenue (current month)
        month_start = datetime(current_year, current_month, 1)
        if current_month == 12:
            month_end = datetime(current_year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = datetime(current_year, current_month + 1, 1) - timedelta(days=1)
        
        monthly_revenue = db.session.query(func.sum(Transaction.amount * 0.1))\
            .filter(Transaction.status == 'completed',
                   Transaction.completion_date.between(month_start, month_end)).scalar() or 0
        
        # Customer growth (year-over-year)
        last_year_start = datetime(current_year - 1, 1, 1)
        last_year_end = datetime(current_year - 1, 12, 31, 23, 59, 59)
        
        customers_this_year = User.query.filter(
            User.role == 'customer',
            User.created_at >= ytd_start
        ).count()
        
        customers_last_year = User.query.filter(
            User.role == 'customer',
            User.created_at.between(last_year_start, last_year_end)
        ).count()
        
        customer_growth = ((customers_this_year - customers_last_year) / customers_last_year * 100) if customers_last_year > 0 else 0
        
        # Merchant growth
        merchants_this_year = User.query.filter(
            User.role == 'merchant',
            User.created_at >= ytd_start
        ).count()
        
        merchants_last_year = User.query.filter(
            User.role == 'merchant',
            User.created_at.between(last_year_start, last_year_end)
        ).count()
        
        merchant_growth = ((merchants_this_year - merchants_last_year) / merchants_last_year * 100) if merchants_last_year > 0 else 0
        
        # Collection rate (year-to-date)
        completed_plans_ytd = InstalmentPlan.query.filter(
            InstalmentPlan.status == 'completed',
            InstalmentPlan.completed_at >= ytd_start
        ).count()
        
        total_plans_ytd = InstalmentPlan.query.filter(
            InstalmentPlan.created_at >= ytd_start
        ).count()
        
        collection_rate = round((completed_plans_ytd / total_plans_ytd * 100) if total_plans_ytd > 0 else 0, 1)
        
        return {
            "ytd_revenue": float(ytd_revenue),
            "monthly_revenue": float(monthly_revenue),
            "customer_growth": round(customer_growth, 1),
            "merchant_growth": round(merchant_growth, 1),
            "collection_rate": collection_rate,
            "active_customers": User.query.filter(User.role == 'customer', User.status.in_(['approved', 'active'])).count(),
            "active_merchants": User.query.filter(User.role == 'merchant', User.status.in_(['approved', 'active'])).count(),
            "total_transactions_ytd": Transaction.query.filter(Transaction.created_at >= ytd_start).count()
        }, 200